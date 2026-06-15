#!/usr/bin/env python3
"""
ChemiCheck 오픈데이터 → SQLite 변환 스크립트
실행: python3 scripts/build_database.py
출력: ChemiCheck/Storage/DummyData/chemicheck.sqlite
"""
import sqlite3, csv, os, sys, re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl 필요: pip3 install openpyxl")
    sys.exit(1)

REPO_ROOT = Path(__file__).parent.parent
DATA_DIR = REPO_ROOT / "open_data"
DB_PATH = REPO_ROOT / "ChemiCheck" / "Storage" / "DummyData" / "chemicheck.sqlite"

# Household-relevant product categories to include (인쇄용 잉크/토너 등 제외)
EXCLUDED_CATEGORIES = {
    "인쇄용 잉크·토너", "문신용 염료", "인주", "공연용 포그액"
}

def risk_from_symptoms(general: str, inhale: str) -> int:
    text = (general or "") + (inhale or "")
    if not text.strip() or text.strip() in ("·자료없음", "자료없음", ""):
        return 1
    if any(k in text for k in ["치명적", "사망", "발암", "폭발", "폐사", "치명"]):
        return 5
    if any(k in text for k in ["독성", "폐부종", "화상", "신장손상", "간손상", "심각한 손상"]):
        return 4
    if any(k in text for k in ["자극", "메스꺼움", "두통", "피부염", "기침", "호흡"]):
        return 3
    return 2

def clean(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    # Remove leading ·
    if s.startswith("·"):
        s = s[1:].strip()
    return s

def truncate(s: str, n: int) -> str:
    return s[:n] if s else ""


def main():
    print(f"DB 출력: {DB_PATH}")
    DB_PATH.unlink(missing_ok=True)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # ──────────────────────────────────────────
    # 1. chemicals 테이블
    # ──────────────────────────────────────────
    c.execute("""CREATE TABLE chemicals (
        id            INTEGER PRIMARY KEY,
        name_kr       TEXT NOT NULL,
        name_en       TEXT,
        cas_number    TEXT,
        risk_level    INTEGER DEFAULT 2,
        symptom_general TEXT,
        symptom_inhale  TEXT,
        symptom_skin    TEXT,
        symptom_eye     TEXT,
        symptom_oral    TEXT
    )""")

    csv_path = DATA_DIR / "기후에너지환경부 화학물질안전원_화학물질안전관리정보_20260513.csv"
    chem_rows = []
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name_kr = clean(row.get("물질명(국문)", ""))
            name_en = clean(row.get("물질명(영문)", ""))
            cas    = clean(row.get("카스번호(CAS 번호)", ""))
            general = clean(row.get("일반증상", ""))
            inhale  = clean(row.get("흡입", ""))
            skin    = clean(row.get("피부", ""))
            eye     = clean(row.get("안구", ""))
            oral    = clean(row.get("경구", ""))
            if not name_kr:
                continue
            risk = risk_from_symptoms(general, inhale)
            chem_rows.append((
                name_kr, name_en, cas, risk,
                truncate(general, 600), truncate(inhale, 400),
                truncate(skin, 300), truncate(eye, 300), truncate(oral, 300),
            ))
    c.executemany(
        "INSERT INTO chemicals (name_kr,name_en,cas_number,risk_level,"
        "symptom_general,symptom_inhale,symptom_skin,symptom_eye,symptom_oral) "
        "VALUES (?,?,?,?,?,?,?,?,?)", chem_rows
    )
    print(f"✓ chemicals: {len(chem_rows)}건")

    # ──────────────────────────────────────────
    # 2. recalls 테이블 (생활화학제품 위반정보)
    # ──────────────────────────────────────────
    c.execute("""CREATE TABLE recalls (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        product_name   TEXT NOT NULL,
        manufacturer   TEXT,
        seller         TEXT,
        action_type    TEXT,
        action_date    TEXT,
        report_number  TEXT,
        legal_basis    TEXT
    )""")

    wb_viol = openpyxl.load_workbook(
        DATA_DIR / "생활화학제품 위반정보_20260607.xlsx", data_only=True
    )
    ws_viol = wb_viol["위반정보"]
    # Row 1: None | Row 2: header | Row 3+: data
    # Cols: 번호(0) 제품명(1) 제조업체명(2) 제조주소(3) 제조전화(4)
    #       판매업체명(5) 판매주소(6) 판매전화(7) 구분(8) 출처(9) 조치일(10) 신고번호(11)
    viol_rows = []
    for row in ws_viol.iter_rows(min_row=3, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        viol_rows.append((
            clean(row[1]),   # product_name
            clean(row[2]),   # manufacturer
            clean(row[5]),   # seller
            clean(row[8]),   # action_type (제조/수입/판매 구분)
            clean(row[10]),  # action_date
            clean(row[11]),  # report_number
            clean(row[9]),   # legal_basis
        ))
    c.executemany(
        "INSERT INTO recalls (product_name,manufacturer,seller,action_type,"
        "action_date,report_number,legal_basis) VALUES (?,?,?,?,?,?,?)", viol_rows
    )
    print(f"✓ recalls: {len(viol_rows)}건")

    # ──────────────────────────────────────────
    # 3. products 테이블
    #    3a. 안전확인대상(신고) — 유효 제품 only, 제외 카테고리 빼고
    # ──────────────────────────────────────────
    c.execute("""CREATE TABLE products (
        id                  INTEGER PRIMARY KEY AUTOINCREMENT,
        product_name        TEXT NOT NULL,
        category            TEXT,
        manufacturer        TEXT,
        registration_number TEXT,
        is_approved         INTEGER DEFAULT 0,
        discloses_ingredients INTEGER DEFAULT 0
    )""")

    # 전성분공개 제품 신고번호 세트
    ingredient_disclosure_numbers = set()
    ingr_csv = DATA_DIR / "한국환경산업기술원_화학제품관리시스템_생활화학제품 전성분 공개 제품_20250811.csv"
    with open(ingr_csv, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            num = row.get("자가검사번호(신고번호)", "").strip()
            if num:
                ingredient_disclosure_numbers.add(num)

    wb_shin = openpyxl.load_workbook(
        DATA_DIR / "안전확인대상 생활화학제품(신고)_20260607.xlsx", data_only=True
    )
    ws_shin = wb_shin.active
    prod_rows = []
    skipped = 0
    for row in ws_shin.iter_rows(min_row=2, values_only=True):
        status   = clean(row[7])
        category = clean(row[3])
        if status != "유효":
            skipped += 1
            continue
        if category in EXCLUDED_CATEGORIES:
            skipped += 1
            continue
        reg_num = clean(row[5])
        prod_rows.append((
            clean(row[1]),  # product_name
            category,
            clean(row[4]),  # manufacturer
            reg_num,
            0,              # is_approved
            1 if reg_num in ingredient_disclosure_numbers else 0,
        ))
    c.executemany(
        "INSERT INTO products (product_name,category,manufacturer,"
        "registration_number,is_approved,discloses_ingredients) "
        "VALUES (?,?,?,?,?,?)", prod_rows
    )
    print(f"✓ products (신고 유효): {len(prod_rows)}건, 제외: {skipped}건")

    # 3b. 안전확인대상(승인)
    wb_appr = openpyxl.load_workbook(
        DATA_DIR / "안전확인대상 생활화학제품(승인)_20260607.xlsx", data_only=True
    )
    ws_appr = wb_appr["승인대상 안전확인대상생활화학제품"]
    appr_rows = []
    for row in ws_appr.iter_rows(min_row=3, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        appr_rows.append((
            clean(row[1]), clean(row[2]), clean(row[3]), clean(row[4]), 1, 0
        ))
    c.executemany(
        "INSERT INTO products (product_name,category,manufacturer,"
        "registration_number,is_approved,discloses_ingredients) "
        "VALUES (?,?,?,?,?,?)", appr_rows
    )
    print(f"✓ products (승인): {len(appr_rows)}건")

    # 3c. 자율안전정보공개제품 (is_approved=2 → voluntary disclosure)
    wb_vol = openpyxl.load_workbook(
        DATA_DIR / "자율안전정보공개제품 목록_20260607.xlsx", data_only=True
    )
    ws_vol = list(wb_vol.worksheets)[0]
    vol_rows = []
    for row in ws_vol.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        vol_rows.append((
            clean(row[2]) if len(row) > 2 else "",  # product_name
            clean(row[3]) if len(row) > 3 else "",  # category
            clean(row[1]),                           # manufacturer
            clean(row[5]) if len(row) > 5 else "",  # reg_number
            2, 1
        ))
    if vol_rows:
        c.executemany(
            "INSERT INTO products (product_name,category,manufacturer,"
            "registration_number,is_approved,discloses_ingredients) "
            "VALUES (?,?,?,?,?,?)", vol_rows
        )
        print(f"✓ products (자율공개): {len(vol_rows)}건")

    # ──────────────────────────────────────────
    # 4. 인덱스 생성
    # ──────────────────────────────────────────
    c.execute("CREATE INDEX idx_chem_name_kr ON chemicals(name_kr)")
    c.execute("CREATE INDEX idx_chem_cas ON chemicals(cas_number)")
    c.execute("CREATE INDEX idx_prod_name ON products(product_name)")
    c.execute("CREATE INDEX idx_prod_cat ON products(category)")
    c.execute("CREATE INDEX idx_recall_product ON recalls(product_name)")

    conn.commit()
    conn.close()

    size_mb = DB_PATH.stat().st_size / 1024 / 1024
    print(f"\n✅ 완료: {DB_PATH}")
    print(f"   파일 크기: {size_mb:.1f} MB")


if __name__ == "__main__":
    main()
